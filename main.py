from tkinter import *
from tkinter.ttk import *
import tksheet
from openpyxl import load_workbook 
import datetime, time
import Download
import Read
import Ratings
import Send

class Application:
    def __init__(self, root):
        self.root = root
        self.root.title("Chess LV")
        self.root.geometry('240x120') 
        
        # Button to download 
        self.download_button = Button(root, text="Download regulations", command=self.download)
        self.download_button.pack()
        
        # Button to display tournaments
        self.display_tournaments_button = Button(root, text="Display tournaments", command=self.display_tournaments)
        self.display_tournaments_button.pack()
        
        # Button to select tournament
        self.select_tournaments_button = Button(root, text="Select tournament", command=self.select_tournament)
        self.select_tournaments_button.pack()

        global state
        # Label to display the last download time
        self.display_label = Label(root, text=f"Last updated time: {state}")
        self.display_label.pack()

        if state == "NONE":
            self.display_tournaments_button.config(state="disabled")
            self.select_tournaments_button.config(state="disabled")

    def download(self):
        self.download_button.config(state="disabled")
        self.display_tournaments_button.config(state="disabled")
        self.select_tournaments_button.config(state="disabled")

        # Update label text to indicate downloading
        self.display_label.config(text="Downloading...")

        # Schedule the download process in the background
        self.root.after(100, self.perform_download)
        
    def perform_download(self):
        Download.download()
        Read.read_regulations()
        TIME = str(datetime.datetime.now())[:19]
        self.display_label.config(text="Last updated time: " + TIME)

        # Enable buttons after download
        self.download_button.config(state="normal")
        self.display_tournaments_button.config(state="normal")
        self.select_tournaments_button.config(state="normal")

        with open("update.txt","w") as f:
            f.write(TIME)
    
    def display_tournaments(self):
        newWindow = Toplevel(self.root)
 
        newWindow.title("Tournaments")

        newWindow.geometry("720x480")
    

        sheet = tksheet.Sheet(newWindow, width=720, height=480)
        sheet.set_column_widths([180, 180, 180, 180])
        sheet.grid()

        workbook = load_workbook("final.xlsx")
        worksheet = workbook.active

        max_row = worksheet.max_row
        headers = [[worksheet[f"{cj}1"].value for cj in ["A" , "B", "D", "E"]]]
        for i in range(max_row, 1, -1):
            tournament = []
            tournament.append(worksheet['A'+str(i)].value)
            tournament.append(worksheet['B'+str(i)].value)
            tournament.append(worksheet['D'+str(i)].value)
            tournament.append(worksheet['E'+str(i)].value)
            headers.append(tournament)

        workbook.close()

        sheet.set_sheet_data([[col for col in line] for line in headers])
        # table enable choices listed below:
        sheet.enable_bindings(("single_select",
                            "row_select",
                            "column_width_resize",
                            "arrowkeys",
                            "right_click_popup_menu",
                            "rc_select",
                            "rc_insert_row",
                            "rc_delete_row",
                            "copy",
                            "cut",
                            "paste",
                            "delete",
                            "undo",
                            "edit_cell"))

    def select_tournament(self):
        newWindow = Toplevel(self.root)
 
        newWindow.title("Select tournament")

        newWindow.geometry("720x480")

        # my_scrollbar = tk.Scrollbar(newWindow, orient=VERTICAL, command=my_canvas.yview)
        # my_scrollbar.pack(side=RIGHT, fill=Y)

        self.send_email_button = Button(newWindow, text="Continue", command=lambda: self.apply(var.get()))
        self.send_email_button.pack()
        self.send_email_button.config(state="disabled")

        def sel():
            self.send_email_button.config(state="normal")

        var = IntVar(newWindow)

        workbook = load_workbook("final.xlsx")
        worksheet = workbook.active

        max_row = worksheet.max_row

        for i in range(max_row, 1, -1):
            R = Radiobutton(newWindow, text=f"{worksheet['A'+str(i)].value} {worksheet['B'+str(i)].value}", variable=var, value=i, command=sel)
            R.pack( anchor = W )

        label = Label(newWindow)
        label.pack()

        workbook.close()

    def apply(self, idx):
        newWindow = Toplevel(self.root)
        newWindow.title("Apply")
        newWindow.geometry("480x240")

        workbook = load_workbook("final.xlsx")
        worksheet = workbook.active

        dates = worksheet['A'+str(idx)].value
        name = worksheet['B'+str(idx)].value
        email = worksheet['D'+str(idx)].value
        event_type = worksheet['E'+str(idx)].value
        
        workbook.close()

        self.show_date = Label(newWindow, text=f"Dates: {dates}")
        self.show_date.pack()

        self.show_name = Label(newWindow, text=f"Tournament name: {name}")
        self.show_name.pack()

        self.show_email = Label(newWindow, text=f"Email: {email}")
        self.show_email.pack()

        self.show_type = Label(newWindow, text=f"Event type: {event_type}")
        self.show_type.pack()
        
        rating = "Unknown"
        if event_type != "Unknown":
            rating = Ratings.rating(event_type)

        self.show_rating = Label(newWindow, text=f"Rating: {rating}")
        self.show_rating.pack()

        if event_type == "Unknown" or email == "Unknown":
            self.error_label = Label(newWindow, text="It is not possible to send email automatically, do it manually")
            self.error_label.pack()
        else:
            self.confirm_email = Entry(newWindow,width = 30)
            self.confirm_email.pack()

            self.confirm_label = Label(newWindow, text="Confirm email")
            self.confirm_label.pack()

            final_email = ""

            def update_email(self):
                nonlocal final_email
                final_email = self.confirm_email.get()
                self.send_email_button.config(state="normal")

            self.update_email_button = Button(newWindow, text="Update email", command=lambda: update_email(self))
            self.update_email_button.pack()

            self.send_email_button = Button(newWindow, text="Send email", command=lambda: self.send_email(final_email, name+" "+dates, rating))
            self.send_email_button.pack()
            self.send_email_button.config(state="disabled")

    def send_email(self, email, tournament_name, rating):
        Send.send_email(email, tournament_name, rating)
        self.confirm_label.config(text="Message is sent") 
        self.update_email_button.config(state="disabled")
        self.send_email_button.config(state="disabled")     

        
if __name__ == "__main__":
    global state
    state = ""
    with open("update.txt","r") as f:
        state = f.read()
    root = Tk()
    clock = Application(root)
    root.mainloop()